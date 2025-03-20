import pandas as pd
import argparse
import datetime
from openpyxl import load_workbook


def logActions(level, short_desc, long_desc):
    dt_object = datetime.datetime.now()
    dt_string = dt_object.strftime("%m/%d/%Y %H:%M:%S")
    prefix = f"{dt_string} - {level}:"
    print(f"{prefix} {short_desc}")
    if long_desc:
        print(f"{prefix} {long_desc}")


def get_existing_mod_worksheets(file_path):
    try:
        wb = load_workbook(file_path, read_only=False)
        if "Mod_TemplateConfigs" in wb.sheetnames:
            old_mod_df = pd.read_excel(file_path, sheet_name="Mod_TemplateConfigs")
        else:
            old_mod_df = pd.DataFrame()
        if "Mod_VolumeConfigs" in wb.sheetnames:
            old_vol_df = pd.read_excel(file_path, sheet_name="Mod_VolumeConfigs")
        else:
            old_vol_df = pd.DataFrame()

        logActions(
            "INF",
            f"Successfully parsed existing Mod worksheets from XLS document ({file_path})",
            None,
        )
        return old_mod_df, old_vol_df

    except Exception as e:
        logActions(
            "ERR",
            f"Failed to parse existing Mod worksheets from XLS document ({file_path})",
            e,
        )


def get_excel_data(file_path):
    try:
        list_df = pd.read_excel(file_path, sheet_name="List")
        drs_df = pd.read_excel(file_path, sheet_name="DRS_Details")
        drs_vols_df = pd.read_excel(file_path, sheet_name="DRS_Vol_Details")

        wb = load_workbook(file_path, read_only=True)
        if "EC2_Details" in wb.sheetnames:
            ec2_df = pd.read_excel(file_path, sheet_name="EC2_Details")
        else:
            ec2_df = pd.DataFrame()
        if "EC2_Vol_Details" in wb.sheetnames:
            ec2_vols_df = pd.read_excel(file_path, sheet_name="EC2_Vol_Details")
        else:
            ec2_vols_df = pd.DataFrame()

        logActions(
            "INF", f"Successfully parsed data from XLS document ({file_path})", None
        )
        return list_df, drs_df, ec2_df, drs_vols_df, ec2_vols_df

    except Exception as e:
        logActions("ERR", f"Failed to parse data from XLS document ({file_path})", e)


def create_comparison_data(
    list_df, drs_df, ec2_df, drs_vols_df, ec2_vols_df, old_mod_df, old_vol_df
):
    gnrl_data = []
    vol_data = []

    for index, list_row in list_df.iterrows():
        try:
            ss_id = list_row["SourceServerID"]
            hostname = list_row["Hostname"]
            instance_id = list_row["OriginInstanceID"]

            if old_mod_df.empty:
                new_launch_state = ""
                new_subnet_id = ""
                new_copy_pip = ""
                new_ips = ""
                new_rightsizing = ""
                new_instance_type = ""
                new_sg_ids = ""
            else:
                new_launch_state = old_mod_df.loc[
                    old_mod_df["OriginInstanceID"] == instance_id, "New_LaunchState"
                ].squeeze()
                new_subnet_id = old_mod_df.loc[
                    old_mod_df["OriginInstanceID"] == instance_id, "New_DrsSubnetID"
                ].squeeze()
                new_copy_pip = old_mod_df.loc[
                    old_mod_df["OriginInstanceID"] == instance_id, "New_CopyPrivateIP"
                ].squeeze()
                new_ips = old_mod_df.loc[
                    old_mod_df["OriginInstanceID"] == instance_id, "New_PrivateIPs"
                ].squeeze()
                new_rightsizing = old_mod_df.loc[
                    old_mod_df["OriginInstanceID"] == instance_id, "New_RightSizing"
                ].squeeze()
                new_instance_type = old_mod_df.loc[
                    old_mod_df["OriginInstanceID"] == instance_id, "New_InstanceType"
                ].squeeze()
                new_sg_ids = old_mod_df.loc[
                    old_mod_df["OriginInstanceID"] == instance_id,
                    "New_SecurityGroupIDs",
                ].squeeze()

            if ec2_df.empty:
                ec2_subnet_name = ""
                ec2_ips = ""
                ec2_instance_type = ""
                ec2_sg_names = ""
            else:
                ec2_subnet_name = ec2_df.loc[
                    ec2_df["InstanceID"] == instance_id, "Subnet_Name"
                ].squeeze()
                ec2_ips = ec2_df.loc[
                    ec2_df["InstanceID"] == instance_id, "PrivateIPs"
                ].squeeze()
                ec2_instance_type = ec2_df.loc[
                    ec2_df["InstanceID"] == instance_id, "InstanceType"
                ].squeeze()
                ec2_sg_names = ec2_df.loc[
                    ec2_df["InstanceID"] == instance_id, "SecurityGroupNames"
                ].squeeze()

            gnrl_data.append(
                {
                    "Hostname": hostname,
                    "SourceServerID": ss_id,
                    "OriginInstanceID": instance_id,
                    "DRS_LaunchState": drs_df.loc[
                        drs_df["OriginInstanceID"] == instance_id, "LaunchState"
                    ].squeeze(),
                    "New_LaunchState": new_launch_state,
                    "EC2_SubnetName": ec2_subnet_name,
                    "DRS_SubnetName": drs_df.loc[
                        drs_df["OriginInstanceID"] == instance_id, "SubnetName"
                    ].squeeze(),
                    "New_DrsSubnetID": new_subnet_id,
                    "CopyPrivateIP": drs_df.loc[
                        drs_df["OriginInstanceID"] == instance_id, "CopyPrivateIP"
                    ].squeeze(),
                    "New_CopyPrivateIP": new_copy_pip,
                    "EC2_PrivateIPs": ec2_ips,
                    "DRS_PrivateIPs": drs_df.loc[
                        drs_df["OriginInstanceID"] == instance_id, "PrivateIPs"
                    ].squeeze(),
                    "New_PrivateIPs": new_ips,
                    "DRS_RightSizing": drs_df.loc[
                        drs_df["OriginInstanceID"] == instance_id, "Rightsizing"
                    ].squeeze(),
                    "New_RightSizing": new_rightsizing,
                    "EC2_InstanceType": ec2_instance_type,
                    "DRS_InstanceType": drs_df.loc[
                        drs_df["OriginInstanceID"] == instance_id, "InstanceType"
                    ].squeeze(),
                    "New_InstanceType": new_instance_type,
                    "EC2_SecurityGroups": ec2_sg_names,
                    "DRS_SecurityGroups": drs_df.loc[
                        drs_df["OriginInstanceID"] == instance_id, "SecurityGroupNames"
                    ].squeeze(),
                    "DRS_SecurityGroupIDs": drs_df.loc[
                        drs_df["OriginInstanceID"] == instance_id, "SecurityGroupIDs"
                    ].squeeze(),
                    "New_SecurityGroupIDs": new_sg_ids,
                }
            )

            drs_vols_filtered_df = drs_vols_df[
                drs_vols_df["OriginInstanceID"] == instance_id
            ].sort_values(by="Size", ascending=True)
            if not ec2_vols_df.empty:
                ec2_vols_filtered_df = ec2_vols_df[
                    ec2_vols_df["InstanceID"] == instance_id
                ].sort_values(by="Size", ascending=True)

            if old_vol_df.empty:
                old_vols_filtered_df = old_vol_df
            else:
                old_vols_filtered_df = old_vol_df[
                    old_vol_df["OriginInstanceID"] == instance_id
                ].sort_values(by="DRS_Size", ascending=True)

            for index, row in drs_vols_filtered_df.iterrows():
                drs_device_name = drs_vols_filtered_df["DeviceName"][index]
                drs_size = drs_vols_filtered_df["Size"][index]
                drs_type = drs_vols_filtered_df["Type"][index]
                drs_iops = drs_vols_filtered_df["IOPS"][index]
                drs_tp = drs_vols_filtered_df["Throughput"][index]
                
                if not ec2_vols_df.empty:
                    ec2_device_name = ec2_vols_filtered_df["DeviceName"][index]
                    ec2_size = ec2_vols_filtered_df["Size"][index]
                    ec2_type = ec2_vols_filtered_df["Type"][index]
                    ec2_iops = ec2_vols_filtered_df["IOPS"][index]
                    ec2_tp = ec2_vols_filtered_df["Throughput"][index]
                else:
                    ec2_device_name = ""
                    ec2_size = ""
                    ec2_type = ""
                    ec2_iops = ""
                    ec2_tp = ""

                if old_vols_filtered_df.empty:
                    new_type = ""
                    new_iops = ""
                    new_throughput = ""
                else:
                    new_type = old_vols_filtered_df["New_Type"][index]
                    new_iops = old_vols_filtered_df["New_IOPS"][index]
                    new_throughput = old_vols_filtered_df["New_Throughput"][index]

                vol_data.append(
                    {
                        "Hostname": hostname,
                        "SourceServerID": ss_id,
                        "OriginInstanceID": instance_id,
                        "EC2_DeviceName": ec2_device_name,
                        "DRS_DeviceName": drs_device_name,
                        "EC2_Size": ec2_size,
                        "DRS_Size": drs_size,
                        "EC2_Type": ec2_type,
                        "DRS_Type": drs_type,
                        "New_Type": new_type,
                        "EC2_IOPS": ec2_iops,
                        "DRS_IOPS": drs_iops,
                        "New_IOPS": new_iops,
                        "EC2_Throughput": ec2_tp,
                        "DRS_Throughput": drs_tp,
                        "New_Throughput": new_throughput,
                    }
                )

            logActions(
                "INF",
                f"Successfully created comparison data for {ss_id} ({hostname})",
                None,
            )
        except Exception as e:
            logActions(
                "ERR", f"Failed to create comparison data for {ss_id} ({hostname})", e
            )

    return gnrl_data, vol_data


def update_workbook(gnrl_data, vol_data, file_path):
    try:
        gnrl_data_df = pd.DataFrame(gnrl_data)
        vol_data_df = pd.DataFrame(vol_data)
        with pd.ExcelWriter(
            file_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            gnrl_data_df.to_excel(writer, sheet_name="Mod_TemplateConfigs", index=False)
            vol_data_df.to_excel(writer, sheet_name="Mod_VolumeConfigs", index=False)

        logActions(
            "INF",
            f"Successfully updated workbook with the comparison data ({file_path})",
            None,
        )

    except Exception as e:
        logActions(
            "ERR",
            f"Failed to update workbook with the comparison data ({file_path})",
            e,
        )


if __name__ == "__main__":
    parser = argparse.ArgumentParser()

    parser.add_argument(
        "--workbook-path", type=str, required=False, help="Path to the XLSX file"
    )
    # Parse the arguments
    args = parser.parse_args()
    file_path = args.workbook_path

    if file_path == None:
        file_path = "DRS_Templates.xlsx"

    old_mod_df, old_vol_df = get_existing_mod_worksheets(file_path)
    list_df, drs_df, ec2_df, drs_vols_df, ec2_vols_df = get_excel_data(file_path)
    gnrl_data, vol_data = create_comparison_data(
        list_df, drs_df, ec2_df, drs_vols_df, ec2_vols_df, old_mod_df, old_vol_df
    )
    update_workbook(gnrl_data, vol_data, file_path)
    logActions("INF", f"Execution finished", None)
